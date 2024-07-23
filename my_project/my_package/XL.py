import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

"""
def switchToNumber(number) :
#my test to pull

    letter = None
    match number :
        case  1 :
            letter = 'A'
        case 2:
            letter = 'B'
        case 3:
            letter = 'C'
        case 4:
            letter = 'D'
        case 5:
            letter = 'E'
        case 6:
            letter = 'F'
        case 7:
            letter = 'G'
        case 8:
            letter = 'H'
        case 9:
            letter = 'I'
        case 10:
            letter = 'J'
    return letter


location =r"C:\selenium\excel_tests.xlsx"
wb = load_workbook(location)
sheet = wb['s1']

# for colidx in sheet.iter_cols(sheet.min_col, sheet.max_col):
#     print(sheet.cell(1,colidx).value)


#for index in range(sheet.max_row):
# print(sheet["A"+str(index+1)].value)








i = 1
cell = switchToNumber(i)
row = sheet.min_row
fcell = sheet.min_column
lcell = sheet.max_column
print(str(fcell) + ' -  ' + str(lcell))
cr = switchToNumber(i) + str(row)

#השדה האחרון עם תוכן
print(sheet.max_row)
print(sheet.min_row)

for val in range (1,sheet.max_row-1,1) :

     print(sheet["A"+ str(row)])
     print(sheet["A"+ str(row)].value)

     row+=1




#השדה הראשון עם תוכן
print(sheet.min_row)
# for x in range(10) :
#     cr = switchToNumber(i) + str(row)
#     print(sheet[switchToNumber(i) + str(row)].value)
#     row+=1
#     i +=1

# sheet['A1'].value = "booo"
# wb.save(location)
# print(sheet['A1'].value)
"""

# תרגיל 1
# location =r"C:\selenium\excel_tests.xlsx"
# wb = load_workbook(location)
# sheet = wb['s1']
#range = sheet['A2' : 'A5']
# i = 2
# for cell in range :
#     for x in cell :
#       age =   input("hello " + x.value + " what is your age? ")
#       sheet['B'+str(i)].value = age
#       i+=1
#---------תרגיל 1 פתרון נוסף
# for index in range(sheet.max_row-1):
#     print(sheet["A"+str(index+2)].value)
#     age = input("hello " + sheet["A"+str(index+2)].value + " what is your age? ")
#     sheet["B"+str(index+2)].value = age
#
# wb.save(location)

# תרגיל 2
# sum = 0
# chemp = None
# max = 0
# for index in range(sheet.max_row-1):
#
#     sum+= int(sheet["B"+str(index+2)].value)
#
#     if int(sheet["B"+str(index+2)].value) > max:
#         max = int(sheet["B"+str(index+2)].value)
#         chemp = sheet["A"+str(index+2)].value
#
#
# print("the avg is : ")
# print(int(sum/(sheet.max_row-1)))
#
# print ("The highest score is of: " + chemp)

# תרגיל 3
# Agroup = 0
# Bgroup = 0
# Cgroup = 0
#
# for index in range(sheet.max_row-1):
#
#     if sheet["B"+str(index+2)].value=="A" :
#         Agroup+=1
#     elif sheet["B"+str(index+2)].value=="B" :
#         Bgroup +=1
#     elif sheet["B"+str(index+2)].value=="C" :
#         Cgroup +=1
#
# print( " in A group we have : " + str(Agroup))
# print( " in B group we have : " + str(Bgroup))
# print( " in C group we have : " + str(Cgroup))

#תרגיל 4
# max = 0
# for index in range(sheet.max_row-1):
#     price = str((sheet["D" + str(index + 2)].value)).replace(',','')
#     if float(price) > max:
#         max = float(price)
#
# print(max)


#*************תרגילים 2 - עם סלניום ****************

#1


# location =r"C:\selenium\excel_tests.xlsx"
# wb = load_workbook(location)
# sheet = wb['basketball']
# driver = webdriver.Chrome()
# driver.maximize_window()
# driver.get("https:#www.google.co.il/search?q=m")
#
# for index in range(sheet.max_row-1):
#      value = sheet["A"+str(index+2)].value
#      driver.find_element(By.NAME, 'q').clear()
#      driver.find_element(By.NAME, 'q').send_keys(value)
#      driver.find_element(By.CLASS_NAME, 'Tg7LZd').click()
#      titlegoogle = driver.find_elements(By.TAG_NAME,"h3")[0].text
#      sheet["B" + str(index + 2)] = titlegoogle
#
# wb.save(location)

#2

# def locat() :
#     user = driver.find_element(By.ID, "uaertName")
#     password = driver.find_element(By.ID, "password")
#     button = driver.find_element(By.ID, "send")
#     err = driver.find_element(By.ID, "err")
#     return user,password,button,err
#
# location =r"C:\selenium\excel_tests.xlsx"
# wb = load_workbook(location)
# sheet = wb['test']
#
# driver = webdriver.Firefox()
# driver.maximize_window()
# driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")


# user = driver.find_element(By.ID,"uaertName")
# password = driver.find_element(By.ID,"password")
# button = driver.find_element(By.ID,"send")
# err = driver.find_element(By.ID,"err")
#  1 צעד

# try :
#     user,password,button, err =  locat()
#     user.clear()
#     password.clear()
#     user.send_keys("admin")
#     password.send_keys("12345")
#     button.click()
#
#     if err.text == "You have entered incorrect details" :
#         sheet["C2"] = "pass"
#
#     else :
#         sheet["C2"] = "fail"
#
#
#
# except Exception as  e :
#     driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")
#     sheet["C2"] = "טסט נכשל"
#
#
# #2 צעד
#
# try :
#     user, password, button, err = locat()
#     user.clear()
#     password.clear()
#     user.send_keys("admin")
#     password.send_keys("pass")
#     button.click()
#     print(err.text)
#     if err.text == "You have entered incorrect details" :
#         sheet["C3"] = "pass"
#
#     else :
#         sheet["C3"] = "fail"
#
#     user.clear()
#     password.clear()
#
#
#
# except Exception as  e :
#     driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")
#     sheet["C3"] = "טסט 2 נכשל"
#
#
# #3 צעד
#
# try :
#     user, password, button, err = locat()
#     user.clear()
#     password.clear()
#     user.send_keys("admin")
#     password.send_keys("")
#     button.click()
#     print(err.text)
#     if err.text == "You have entered incorrect details" :
#         sheet["C4"] = "pass"
#
#     else :
#         sheet["C4"] = "fail"
#
#
#
#
#
# except Exception as  e :
#     driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")
#     sheet["C4"] = "טסט 3 נכשל"
#
#
#
# #4 צעד
#
# try :
#     user, password, button, err = locat()
#     user.clear()
#     password.clear()
#     user.send_keys("admin")
#     password.send_keys("admin")
#     button.click()
#     time.sleep(3)
#
#     if driver.find_element(By.TAG_NAME,'h1').text == "You've logged in successfully" :
#         sheet["C5"] = "pass"
#
#     else :
#         sheet["C5"] = "FAIL"
#
# except Exception as  e :
#      driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")
#      sheet["C5"] = "טסטס 4 נכשל"
#      print(e)
#
# wb.save(location)

#***************************************3
location =r"C:\selenium\excel_tests.xlsx"
wb = load_workbook(location)
sheet = wb['144']
sheet.cell(row=2, column=2).value = "links"
# driver = webdriver.Chrome()
# driver.maximize_window()


#driver.get("https://www.b144.co.il/clients/%D7%90/")

# for i in range(1, 4):
#     list_elements = driver.find_elements(By.XPATH, "//div[@class='fields-index-con cat-items']//a")
#     charsList = driver.find_elements(By.XPATH, "//div[@id='_Letters_PnlControl']//a")
#     for j in range(11):
#         links = list_elements[j].text
#         print(links)
#         sheet.cell(row=j + 1, column=i + 1).value = links
#
#     print()  # להדפיס שורה ריקה
#     charsList[i].click()
#     time.sleep(1)  # השהיה כדי לאפשר טעינת דף

wb.save(location)  # שמירת הקובץ
wb.close()

#
# driver = webdriver.Firefox()
# driver.maximize_window()
#

input("x ")
