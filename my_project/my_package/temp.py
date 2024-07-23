import csv
import time
import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
# from selenium import webdriver
# driver = webdriver.Firefox()
# driver.maximize_window()
# driver.get("http://danielauto.net/practice/loginPage/loginpageTest.html")
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

#פתיחה עם חוסם פירסומות
# path = r'C:\Users\Daniel\AppData\Local\Google\Chrome\User Data\Default\Extensions\cjpalhdlnbpafiamejdnhcphjbkeiagm\1.58.0_0'
# options = Options()
# options.add_argument("load-extension=" + path)
# driver = webdriver.Chrome(options=options)
# driver.maximize_window()
# driver.get("https://simania.co.il/")
#
from selenium.webdriver.common.keys import Keys
driver = webdriver.Chrome()
driver.maximize_window()

#driver.get("https://doovdevanim.pajonos.com/torah/%D7%A1%D7%A4%D7%A8-%D7%91%D7%A8%D7%90%D7%A9%D7%99%D7%AA/%D7%A4%D7%A8%D7%A9%D7%AA-%D7%97%D7%99%D7%99-%D7%A9%D7%A8%D7%94/")
location =r"C:\Users\Daniel\Documents\doov.xlsx"
wb = load_workbook(location)

sheet = wb['moed']
sheet2 = wb['link']

lista = driver.find_elements(By.TAG_NAME,'h2')
count = 1
for x in range (63,80,1) :

       driver.get(sheet2.cell(row= x, column= 1).value)
       lista = driver.find_elements(By.TAG_NAME, 'h2')
       for i , link  in enumerate(lista) :
         link.click()
         time.sleep(1)
         sheet.cell(row=  count, column= 1).value = driver.find_element(By.XPATH,"//div[@class=\"body\"]//h1").text
         print(driver.find_element(By.XPATH,"//div[@class=\"body\"]//h1").text)
         print(sheet.cell(row=  count, column= 1).value)
         sheet.cell(row= count, column=2).value = driver.find_element(By.XPATH, "//ul[@class = 'breadcrumbs']//li[2]").text
         sheet.cell(row= count, column=3).value = driver.find_element(By.XPATH, "//ul[@class = 'breadcrumbs']//li[3]").text

         source_element = driver.find_element(By.CSS_SELECTOR,"div.metadata > span:nth-of-type(3)")
         source_text = driver.execute_script('return arguments[0].nextSibling.textContent;', source_element).strip()
         sheet.cell(row= count, column=4).value = source_text
         sheet.cell(row= count, column=5).value = driver.find_element(By.XPATH, '//div[@class=\'content\']').text
         wb.save(location)  # שמירת הקובץ
         wb.close()
         driver.back()
         time.sleep(1)
         count+=1



