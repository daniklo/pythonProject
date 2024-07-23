import csv
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl.reader.excel import load_workbook

location =r"C:\selenium\excel_tests.xlsx"
wb = load_workbook(location)
sheet = wb['basketball']

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.google.com/")

for index in range(sheet.max_row-1):
    value = sheet["A"+str(index+2)].value
    driver.find_element(By.NAME, "q").clear()
    driver.find_element(By.NAME, "q").send_keys(value + Keys.ENTER)
    #driver.find_element(By.NAME, "q").send_keys(Keys.ENTER)



    time.sleep(1)
    titlegoogle = driver.find_elements(By.TAG_NAME, "h3")[0].text
    sheet["B" + str(index +2)] = titlegoogle
    print(titlegoogle)